from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog
import sys
from PyQt5.QtGui import QWheelEvent
from PyQt5 import QtGui
from PyQt5.QtWidgets import QWidget
from PyQt5.QtGui import QPixmap,QIcon
from PyQt5.QtCore import Qt, QTimer, PYQT_VERSION_STR
from PyQt5 import QtWidgets, QtGui, QtCore
import cv2
import os
from os.path import expanduser
import openpyxl 
import pandas as pd
from datetime import datetime
import sys
from PyQt5.QtWidgets import QApplication
import numpy as np
import cv2
import matplotlib.pyplot as plt 
import time
import platform
os.environ["SDL_JOYSTICK_ALLOW_BACKGROUND_EVENTS"] = "1"
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
try:
    import EasyPySpin
except Exception:
    pass

from classes.rrt_star_class import RrtStar
from classes.rrt_class import RRT
from classes.tracker_class import VideoThread
from classes.gui_widgets import Ui_MainWindow
from classes.arduino_class import ArduinoHandler
from classes.robot_class import Robot
from classes.record_class import RecordThread
from classes.algorithm_class import algorithm
from classes.generate_data_circles import gen_data
from classes.generate_data_infinity import gen_data2
from classes.Learning_module_2d import LearningModule



class MainWindow(QtWidgets.QMainWindow):
    positionChanged = QtCore.pyqtSignal(QtCore.QPoint)

    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent=parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        
        
        #self.showMaximized()

        #resize some widgets to fit the screen better
        screen  = QtWidgets.QDesktopWidget().screenGeometry(-1)
        
        self.window_width = screen.width()
        self.window_height = screen.height()
        self.resize(self.window_width, self.window_height)
        self.display_width = self.window_width# self.ui.frameGeometry().width()

        self.displayheightratio = 0.79
        self.framesliderheightratio = 0.031
        self.textheightratio = .129
        self.tabheightratio = 0.925
        self.tabheightratio = 0.925
        
        self.aspectratio = 1041/801
        self.resize_widgets()

    
      

        self.new_dir_path = r"D:\mpc\videos"
        if not os.path.exists(self.new_dir_path):
            os.makedirs(self.new_dir_path)

        #connect to arduino
        if "mac" in platform.platform():
            self.tbprint("Detected OS: macos")
            PORT = "/dev/cu.usbmodem11301"
           
        elif "Linux" in platform.platform():
            self.tbprint("Detected OS: Linux")
            PORT = "/dev/ttyACM0"

        elif "Windows" in platform.platform():
            self.tbprint("Detected OS:  Windows")
            PORT = "COM4"
        else:
            self.tbprint("undetected operating system")
            PORT = None
        
        self.arduino = ArduinoHandler(self.tbprint)
        self.arduino.connect(PORT)
        self.algorithm = algorithm()
        self.cycles_gen_data = 3
        self.generate_data = gen_data(self.cycles_gen_data)
        self.generate_data2 = gen_data2()
        self.calibration_coord = [self.algorithm.init_point_x, self.algorithm.init_point_y]

        self.GP = LearningModule(self.cycles_gen_data)

        self.zoom_x, self.zoom_y, self.zoomscale, self.scrollamount = 1,0,0,0
        self.croppedresult = None
        self.currentframe = None
        self.frame_number = 0
        self.robots = []
        self.videopath = 0
        self.cap = None
        self.tracker = None
        self.recorder = None

        #self.dataset_GP = []  #data from generate data function 

        self.save_status = False
        self.output_workbook = None
        

        
     
        self.ui.trackbutton.clicked.connect(self.track)
        self.ui.maskbutton.clicked.connect(self.showmask)
        self.ui.maskinvert_checkBox.toggled.connect(self.invertmaskcommand)
    
        self.ui.robotmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.robotcroplengthbox.valueChanged.connect(self.get_slider_vals)

        self.ui.cellmasklowerbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskupperbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskdilationbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellmaskblurbox.valueChanged.connect(self.get_slider_vals)
        self.ui.cellcroplengthbox.valueChanged.connect(self.get_slider_vals)
      


        #self.ui.savedatabutton.clicked.connect(self.savedata)
        self.ui.VideoFeedLabel.installEventFilter(self)
        self.ui.recordbutton.clicked.connect(self.recordfunction_class)
        self.ui.resetdefaultbutton.clicked.connect(self.resetparams)
        self.ui.objectivebox.valueChanged.connect(self.get_objective)
        self.ui.exposurebox.valueChanged.connect(self.get_exposure)
        self.ui.croppedmasktoggle.clicked.connect(self.showcroppedoriginal)
        self.ui.croppedrecordbutton.clicked.connect(self.croppedrecordfunction)
        #self.ui.import_excel_actions.clicked.connect(self.read_excel_actions)
        #self.ui.apply_actions.clicked.connect(self.apply_excel_actions)
        self.ui.generate_data_button.clicked.connect(self.generate_data_function)
        self.ui.run_algo.clicked.connect(self.run_algorithm)
        self.ui.calibrate_button.clicked.connect(self.go_to_start)
        self.ui.Trainbutton.clicked.connect(self.train_function)
        self.ui.reset_paths.toggled.connect(self.reset_path_function)
        self.ui.gen_data2_button.clicked.connect(self.generate_data2_function)
        self.ui.train_data2_button.clicked.connect(self.train_function2)


        #readomg excel file variables        
        self.excel_file_name = None
        self.excel_actions_df = None
        self.excel_actions_status = False

        self.algorithm_status = False
        self.calibrate_status = False
        self.generate_data_status = False
        self.generate_data_status2 = False
        self.train_status = False
        self.train_status2 = False


    def reset_path_function(self):
        if len(self.tracker.robot_list) > 0:
            self.tracker.robot_list[-1].trajectory.clear()
        
            
    
    def generate_data_function(self):
        if self.ui.generate_data_button.isChecked():
            
            self.generate_data_status = True
            self.ui.generate_data_button.setText("Stop")
        else:

            self.generate_data_status = False
            self.ui.generate_data_button.setText("Gen Data")
            self.generate_data.reset(self.cycles_gen_data)


    def train_function(self):
        if self.ui.Trainbutton.isChecked():
            self.train_status = True
            self.ui.Trainbutton.setText("Stop")
            print("this will only print once")
            dataset =  np.load('datasetGP.npy')
            self.GP.read_data_action(dataset, self.tracker.objective)
       
            self.GP.estimate_a0(0)
        else:

            self.train_status = False
            self.ui.Trainbutton.setText("Train")


    def generate_data2_function(self):
        if self.ui.gen_data2_button.isChecked():
            
            self.generate_data_status2 = True
            self.ui.gen_data2_button.setText("Stop")
        else:

            self.generate_data_status2 = False
            self.ui.gen_data2_button.setText("Gen Data 2")
            
            self.generate_data2.reset()


    def train_function2(self):
        if self.ui.train_data2_button.isChecked():
            self.train_status2 = True
            self.ui.train_data2_button.setText("Stop")
            print("this will only print once")
            dataset2 =  np.load('datasetGP2.npy')
            dataset1 =  np.load('datasetGP.npy')
            self.GP.read_data_action(dataset1, self.tracker.objective)
            self.GP.read_data_action2(dataset2, self.tracker.objective)
       

            print('reached here 1')
            self.GP.estimate_a0(3)
           
        else:

            self.train_status2 = False
            self.ui.train_data2_button.setText("Train 2")

    


    




    def update_image(self, frame, cell_mask, robot_list):
  
        self.cell_mask = cell_mask
        alpha = 0
        freq = 0
        """Updates the image_label with a new opencv image"""
        
        #step 1: generate data 1(circles)
        if self.generate_data_status == True and self.generate_data.run_calibration_status == True: 
            
            if len(robot_list) > 0:
                Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq = self.generate_data.run_calibration_circles(robot_list)

                self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)
            
        

        elif self.generate_data_status == True and self.generate_data.run_calibration_status == False:
            if len(robot_list) > 0:
            
                Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq = self.generate_data.run_circles()
                self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)
                if self.generate_data.reading_actions == True:
                    #in here save data
                    time = robot_list[-1].times[-1]
                    px = robot_list[-1].position_list[-1][0]
                    py = robot_list[-1].position_list[-1][1]
                    vx = robot_list[-1].velocity_list[-1][0]
                    vy = robot_list[-1].velocity_list[-1][1]
                    alpha = alpha
                    freq = freq

                    #save current state to dataset
                    self.generate_data.dataset_GP.append([time, px,py,vx,vy, alpha, freq])
                
                if self.generate_data.reading_completed :
                    print('data size =', len(self.generate_data.dataset_GP))
                    np.save('datasetGP.npy', np.array(self.generate_data.dataset_GP))

        #step 3: generate data 2 (infinity)
        elif self.generate_data_status2 == True and self.generate_data2.run_calibration_status == True: 
            
            if len(robot_list) > 0:
                Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq = self.generate_data2.run_calibration_infinity(robot_list)

                self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)
            
        

        elif self.generate_data_status2 == True and self.generate_data2.run_calibration_status == False:
            if len(robot_list) > 0:
            
                frame, Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq = self.generate_data2.run_infinity(robot_list, frame)
                self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)
                

                #frame, Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq = self.algorithm.run(robot_list, frame)
                #self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)


                if self.generate_data2.reading_actions == True:
                    #in here save data
                    time = robot_list[-1].times[-1]
                    px = robot_list[-1].position_list[-1][0]
                    py = robot_list[-1].position_list[-1][1]
                    vx = robot_list[-1].velocity_list[-1][0]
                    vy = robot_list[-1].velocity_list[-1][1]
                    alpha = alpha
                    freq = freq

                    #save current state to dataset
                    self.generate_data2.dataset_GP2.append([time, px,py,vx,vy, alpha, freq])
                    
                if self.generate_data2.reading_completed:
                    print('data size2 =', len(self.generate_data2.dataset_GP2))
                    np.save('datasetGP2.npy', np.array(self.generate_data2.dataset_GP2))
                    



        
        #step 3
        elif self.calibrate_status == True:
            if len(robot_list) > 0:
                curernt_pos = robot_list[-1].position_list[-1] #the most recent position at the time of clicking run algo
                
                #print(curernt_pos)

                direction_vec = [self.calibration_coord[0] - curernt_pos[0], self.calibration_coord[1] - curernt_pos[1]]
                error = np.sqrt(direction_vec[0] ** 2 + direction_vec[1] ** 2)
                alpha = np.arctan2(-direction_vec[1], direction_vec[0]) - np.pi/2
                freq = 10
                if error < 5:
                    self.arduino.send(0,0,0,0,0,0,0,0,0)
                else:
                    self.arduino.send(0,0,0,alpha,np.pi/2,freq,np.pi/2,0,0)
        
        


        #step 4    
        elif self.algorithm_status == True:
            if len(robot_list) > 0:
                frame, Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq = self.algorithm.run(robot_list, frame)
                self.arduino.send(Bx,By,Bz,alpha,gamma,freq,psi,gradient,acoustic_freq)




        elif self.excel_actions_status == True and self.excel_actions_df is not None:
            
            self.actions_counter +=1

            if self.actions_counter < self.excel_actions_df['Frame'].iloc[-1]:
                filtered_row = self.excel_actions_df[self.excel_actions_df['Frame'] == self.actions_counter]
            
                Bx = float(filtered_row["Bx"])
                By = float(filtered_row["By"])
                Bz = float(filtered_row["Bz"])
                alpha = float(filtered_row["Alpha"])
                gamma = float(filtered_row["Gamma"])
                freq = float(filtered_row["Rolling Frequency"])
                psi = float(filtered_row["Psi"])
                gradient = float(filtered_row["Gradient"])
                acoustic_freq = float(filtered_row["Acoustic Frequency"])
                self.arduino.send(Bx, By, Bz, alpha, gamma, freq, psi, gradient, acoustic_freq)
            
            else:
                self.excel_actions_status = False
                self.ui.apply_actions.setText("Apply")
                self.ui.apply_actions.setChecked(False)
                self.arduino.send(0,0,0,0,0,0,0,0,0)
                alpha = 0
                freq = 0
        else:
            self.arduino.send(0,0,0,0,0,0,0,0,0)  #zeros everything
            alpha = 0
            freq = 0
            
        
        
        #DEFINE CURRENT ROBOT PARAMS TO A LIST
       
        if len(robot_list) > 0:
            self.robots = []
            for bot in robot_list:
                currentbot_params = [bot.frame_list[-1],
                                     bot.times[-1],
                                     alpha,
                                     freq,
                                     bot.position_list[-1][0],bot.position_list[-1][1], 
                                     bot.velocity_list[-1][0], bot.velocity_list[-1][1],bot.velocity_list[-1][2],
                                     bot.blur_list[-1],
                                     bot.area_list[-1],
                                     bot.avg_area,
                                     bot.cropped_frame[-1][0],bot.cropped_frame[-1][1],bot.cropped_frame[-1][2],bot.cropped_frame[-1][3],
                                     bot.stuck_status_list[-1],
                                     bot.trajectory,
                                    ]
                
                self.robots.append(currentbot_params)
        
        #IF SAVE STATUS THEN CONTINOUSLY SAVE THE CURRENT ROBOT PARAMS AND MAGNETIC FIELD PARAMS TO AN EXCEL ROWS
        if self.save_status == True:
            for (sheet, bot) in zip(self.robot_params_sheets,self.robots):
                sheet.append(bot[:-1])



        #HANDLE THE FRAME CAPTURE
        
        

        frame = self.handle_zoom(frame)

        self.currentframe = frame
    
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
      
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(self.display_width, self.display_height, Qt.KeepAspectRatio)
        qt_img = QPixmap.fromImage(p)

       
        self.ui.VideoFeedLabel.setPixmap(qt_img)
        
        

    def go_to_start(self):
        if self.ui.calibrate_button.isChecked():
            self.calibrate_status = True
            self.ui.calibrate_button.setText("Stop")
            self.algorithm.load_GP()
            print("loading GP")
        else:

            #when I click stop, it stops the calibration
            #self.algorithm.generate_traj(self.tracker.robot_list)
            self.calibrate_status = False
            self.ui.calibrate_button.setText("Calibrate")

        


    
      


    def run_algorithm(self):
        if self.ui.run_algo.isChecked():
            self.algorithm_status = True
            self.ui.run_algo.setText("Stop")
        else:

            self.algorithm_status = False
            self.ui.run_algo.setText("Apply Algo")
            self.algorithm.reset()



    def read_excel_actions(self):
        options = QFileDialog.Options()
        self.excel_file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)", options=options)
        if self.excel_file_name:
            self.excel_actions_df = pd.read_excel(self.excel_file_name)
            
        
    def apply_excel_actions(self):
        if self.ui.apply_actions.isChecked():
            self.excel_actions_status = True
            self.actions_counter = 0
            self.ui.apply_actions.setText("Stop")
        else:
            self.excel_actions_status = False
            self.ui.apply_actions.setText("Apply")
            self.arduino.send(0,0,0,0,0,0,0,0,0)

    
      


    def start_data_record(self):
        self.output_workbook = openpyxl.Workbook()
            

        #create sheet for robot data
        self.robot_params_sheets = []
        for i in range(len(self.robots)):
            robot_sheet = self.output_workbook.create_sheet(title= "Robot {}".format(i+1))
            robot_sheet.append(["Frame","Times","Alpha","Freq","Pos X", "Pos Y", "Vel X", "Vel Y", "Vel Mag", "Blur", "Area", "Avg Area", "Cropped X","Cropped Y","Cropped W","Cropped H","Stuck?","Path X", "Path Y"])
            self.robot_params_sheets.append(robot_sheet)
        

        #tell update_actions function to start appending data to the sheets
        self.save_status = True



    def stop_data_record(self):
        #tell update_actions function to stop appending data to the sheets
        self.save_status = False
        file_path  = os.path.join(self.new_dir_path, self.date+".xlsx")
        
        #add trajectory to file after the fact
        if self.output_workbook is not None:
            if len((self.robot_params_sheets)) > 0:
                try:
                    for i in range(len((self.robot_params_sheets))):
                        for idx,(x,y) in enumerate(self.robots[i][-1]):
                            self.robot_params_sheets[i].cell(row=idx+2, column=18).value = x
                            self.robot_params_sheets[i].cell(row=idx+2, column=19).value = y
                except Exception:
                    pass
       
            #save and close workbook
            self.output_workbook.remove(self.output_workbook["Sheet"])
            self.output_workbook.save(file_path)

            self.output_workbook.close()
            self.output_workbook = None

    
    def savedata(self):
        if self.ui.savedatabutton.isChecked():
            self.ui.savedatabutton.setText("Stop")
            self.start_data_record()
        else:
            self.ui.savedatabutton.setText("Save Data")
            self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
            self.stop_data_record()
   
    def tbprint(self, text):
        #print to textbox
        self.ui.plainTextEdit.appendPlainText("$ "+ text)
    

    def convert_coords(self,pos):
        #need a way to convert the video position of mouse to the actually coordinate in the window
        newx = int(pos.x() * (self.video_width / self.display_width)) 
        newy = int(pos.y() * (self.video_height / self.display_height))
        return newx, newy

    def eventFilter(self, object, event):
        if object is self.ui.VideoFeedLabel: 
            if self.tracker is not None:
                if event.type() == QtCore.QEvent.MouseButtonPress:   
                    if event.buttons() == QtCore.Qt.LeftButton:
                        newx, newy = self.convert_coords(event.pos())
                        #generate original bounding box
                        
                 
                        x_1 = int(newx - self.ui.robotcroplengthbox.value()  / 2)
                        y_1 = int(newy - self.ui.robotcroplengthbox.value()  / 2)
                        w = self.ui.robotcroplengthbox.value()
                        h = self.ui.robotcroplengthbox.value()

                        robot = Robot()  # create robot instance
                        robot.add_frame(self.frame_number)
                        robot.add_time(0)
                        robot.add_position([newx,newy])
                        robot.add_velocity([0,0,0])
                        robot.add_crop([x_1, y_1, w, h])
                        robot.add_area(0)
                        robot.add_blur(0)
                        robot.add_stuck_status(0)
                        robot.crop_length = self.ui.robotcroplengthbox.value()
                        self.tracker.robot_list.append(robot) #this has to include tracker.robot_list because I need to add it to that class
                        
                    
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = True
                        newx, newy = self.convert_coords(event.pos())
                        if len(self.tracker.robot_list) > 0:
                            startpos = self.tracker.robot_list[-1].position_list[-1]
                            endpos = [newx, newy]
                     
                            if self.ui.drawingcheckbox.isChecked():
                                self.tracker.robot_list[-1].add_trajectory(startpos)
                            
                            elif self.ui.RRTcheckbox.isChecked():
                                step_size = 50
                                pathplanner = RRT(self.cell_mask, startpos, endpos, step_size)
                
                                trajectory = pathplanner.run()
                                trajectory.append(endpos)    
                            
                                #record robot list trajectory
                                self.tracker.robot_list[-1].trajectory= trajectory


                            elif self.ui.RRTstarcheckbox.isChecked():
                                rrt_star = RrtStar(img = self.cell_mask, x_start = startpos, x_goal=endpos, step_len=50,
                                         goal_sample_rate=.1, search_radius=2, iter_max=3000,plotting_flag=True)
                                
                                self.tracker.robot_list[-1].trajectory = rrt_star.planning()
                            
                
                    if event.buttons() == QtCore.Qt.MiddleButton: 
                        del self.tracker.robot_list[:]
                        del self.robots[:]
            
                       
                    
                            
                elif event.type() == QtCore.QEvent.MouseMove:
                    self.zoom_x, self.zoom_y = self.convert_coords(event.pos())

                    if event.buttons() == QtCore.Qt.RightButton:
                        if self.drawing == True:
                            if len(self.tracker.robot_list)>0:
                                if self.ui.drawingcheckbox.isChecked():
                               
                                    newx, newy = self.convert_coords(event.pos())
                                    
                                    self.tracker.robot_list[-1].add_trajectory([newx, newy])
                  
                                                               
                
                elif event.type() == QtCore.QEvent.MouseButtonRelease:
                    if event.buttons() == QtCore.Qt.RightButton: 
                        self.drawing = False
                        
                if event.type() ==  QtCore.QEvent.Wheel:
                    steps = event.angleDelta().y() 
                    
                    self.scrollamount += (steps and steps / abs(steps/0.5))
                    self.scrollamount = max(min(self.scrollamount,20.0),1.0)
                    self.zoomscale = self.scrollamount

        
        return super().eventFilter(object, event)
            


    

    def update_croppedimage(self, frame, recoreded_frame):
        """Updates the cropped image_label with a new cropped opencv image"""
        
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(310, 310, Qt.KeepAspectRatio)
        qt_cimg = QPixmap.fromImage(p)
        self.ui.CroppedVideoFeedLabel.setPixmap(qt_cimg)
        
        #recored the robots suroundings
        if self.croppedresult is not None:
            self.croppedresult.write(recoreded_frame)

    

    def croppedrecordfunction(self):
        if self.cap is not None:
            if self.ui.croppedrecordbutton.isChecked():
                self.ui.croppedrecordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
                file_path  = os.path.join(self.new_dir_path, self.date+".mp4")
                self.croppedresult = cv2.VideoWriter(
                    file_path,
                    cv2.VideoWriter_fourcc(*"mp4v"),
                    int(self.videofps),    
                    (200, 200), ) 
                #start recording magnetic field and tracking data
                self.start_data_record()
            
            else:
                self.ui.croppedrecordbutton.setText("Record")
                if self.croppedresult is not None:
                    self.croppedresult.release()
                    self.croppedresult = None
                    self.tbprint("End Record, Data Saved")
                #stop and save the data when the record is over.
                self.stop_data_record()
    
         
    def recordfunction_class(self):
        if self.cap is not None:
            if self.ui.recordbutton.isChecked():
                self.date = datetime.now().strftime('%Y.%m.%d-%H.%M.%S')
                self.recorder = RecordThread(self, self.date)
                self.recorder.recordstatus = True
                self.recorder.start()
                self.ui.recordbutton.setText("Stop")
                self.tbprint("Start Record")
                self.start_data_record()
                
            else:
                self.recorder.stop()
                self.ui.recordbutton.setText("Record")
                self.tbprint("End Record, Data Saved")
                self.stop_data_record()


    
    def setFile(self):
        if self.videopath == 0:
            try:
                self.cap  = EasyPySpin.VideoCapture(0)
                self.cap.set(cv2.CAP_PROP_AUTO_WB, True)
                self.cap.set(cv2.CAP_PROP_FPS, 10)
                print("flir cam detected")
            except Exception:
                self.cap  = cv2.VideoCapture(0) 
                self.tbprint("No EasyPySpin Camera Available")
          
        
        self.video_width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        self.video_height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.videofps = int(self.cap.get(cv2.CAP_PROP_FPS))
        self.tbprint("Width: {}  --  Height: {}  --  Fps: {}".format(self.video_width,self.video_height,self.videofps))

        self.aspectratio = (self.video_width / self.video_height)

        self.resize_widgets()        

        if self.videopath == 0:
            self.ui.robotsizeunitslabel.setText("um")
            self.ui.robotvelocityunitslabel.setText("um/s")
        

        self.ui.VideoFeedLabel.setPixmap(QtGui.QPixmap())
        


    def track(self):
        if self.videopath is not None:
            if self.ui.trackbutton.isChecked():
                self.setFile()
                    
                self.tracker = VideoThread(self)
                self.tracker.change_pixmap_signal.connect(self.update_image)
                self.tracker.cropped_frame_signal.connect(self.update_croppedimage)
                self.tracker.start()

                self.ui.trackbutton.setText("Camera Off")
                self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
                self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(0, 255, 0); ")
        
                
            else:
                self.ui.VideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
                self.ui.CroppedVideoFeedLabel.setStyleSheet("background-color: rgb(0,0,0); border:2px solid rgb(255, 0, 0); ")
        
                if self.tracker is not None:
                    self.ui.trackbutton.setText("Camera On")
                    self.tracker.stop()
            

                    #reset mask button
                    self.tracker.mask_flag = False
                    self.ui.maskbutton.setText("Mask")
                    self.ui.maskbutton.setChecked(False)

        
                    #zero arduino commands
                    del self.tracker.robot_list[:]


            

    def showmask(self):
        if self.tracker is not None:
            if self.ui.maskbutton.isChecked():
                self.ui.maskbutton.setText("Original")
                self.tracker.mask_flag = True
            else:
                self.ui.maskbutton.setText("Mask")
                self.tracker.mask_flag = False
    
    def showcroppedoriginal(self):
        if self.tracker is not None:
            if self.ui.croppedmasktoggle.isChecked():
                self.ui.croppedmasktoggle.setText("Mask")
                self.tracker.croppedmask_flag = False
            else:
                self.ui.croppedmasktoggle.setText("Original")
                self.tracker.croppedmask_flag = True


         
    def get_objective(self):
        if self.tracker is not None:
            self.tracker.objective = self.ui.objectivebox.value()

    def get_exposure(self):
        if self.tracker is not None:
            self.tracker.exposure = self.ui.exposurebox.value()
            
    
    def invertmaskcommand(self):
        if self.tracker is not None:
            self.ui.maskinvert_checkBox.setText("Invert Mask: " + str(self.ui.maskinvert_checkBox.isChecked()))
            self.tracker.maskinvert = self.ui.maskinvert_checkBox.isChecked()


    
    
    
    

    def get_slider_vals(self):
        #alpha = self.ui.alphaspinBox.value()
        
        robotlower = self.ui.robotmasklowerbox.value() 
        robotupper = self.ui.robotmaskupperbox.value()
        robotdilation = self.ui.robotmaskdilationbox.value() 
        robotmaskblur = self.ui.robotmaskblurbox.value()
        robotcrop_length = self.ui.robotcroplengthbox.value()

        celllower = self.ui.cellmasklowerbox.value() 
        cellupper = self.ui.cellmaskupperbox.value()
        celldilation = self.ui.cellmaskdilationbox.value() 
        cellmaskblur = self.ui.cellmaskblurbox.value()
        cellcrop_length = self.ui.cellcroplengthbox.value()


        if self.tracker is not None: 

            self.tracker.robot_mask_lower = robotlower
            self.tracker.robot_mask_upper = robotupper
            self.tracker.robot_mask_dilation = robotdilation
            self.tracker.robot_mask_blur = robotmaskblur
            self.tracker.robot_crop_length = robotcrop_length

            self.tracker.cell_mask_lower = celllower
            self.tracker.cell_mask_upper = cellupper
            self.tracker.cell_mask_dilation = celldilation
            self.tracker.cell_mask_blur = cellmaskblur
            self.tracker.cell_crop_length = cellcrop_length



         
        
    def resetparams(self):
        self.ui.robotmasklowerbox.setValue(0)
        self.ui.robotmaskupperbox.setValue(123)
        self.ui.robotmaskdilationbox.setValue(3)
        self.ui.robotmaskblurbox.setValue(6)
        self.ui.robotcroplengthbox.setValue(90)
        self.ui.objectivebox.setValue(10)
        self.ui.exposurebox.setValue(5000)

        self.ui.cellmasklowerbox.setValue(0)
        self.ui.cellmaskupperbox.setValue(128)
        self.ui.cellmaskdilationbox.setValue(0)
        self.ui.cellmaskblurbox.setValue(0)
        self.ui.cellcroplengthbox.setValue(40)
        

    def resizeEvent(self, event):
        windowsize = event.size()
        self.window_width = windowsize.width()
        self.window_height = windowsize.height()
        self.resize_widgets()
 
    def resize_widgets(self):
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio
        self.display_height = int(self.window_height*self.displayheightratio) #keep this fixed, changed the width dpending on the aspect ratio
        self.framesliderheight = int(self.window_height*self.framesliderheightratio)
        self.textheight = int(self.window_height*self.textheightratio)
        self.tabheight = self.window_height*self.tabheightratio

        self.display_width = int(self.display_height * self.aspectratio)

        self.ui.VideoFeedLabel.setGeometry(QtCore.QRect(10,  5,                       self.display_width,     self.display_height))
        self.ui.frameslider.setGeometry(QtCore.QRect(10,    self.display_height+12,   self.display_width,     self.framesliderheight))
        self.ui.plainTextEdit.setGeometry(QtCore.QRect(10,  self.display_height+20+self.framesliderheight,   self.display_width,     self.textheight))

        #self.ui.tabWidget.setGeometry(QtCore.QRect(12,  6,  260 ,     self.tabheight))

    def handle_zoom(self, frame):
        
        if self.zoomscale > 1:
            x = self.zoom_x
            y = self.zoom_y
            w = 300
            h = 300
            angle = 0
            
            # step 1: cropped a frame around the coord you wont to zoom into
            if y-w < 0 and x-h < 0:
                zoomedframe = frame[0:y+h , 0:x+w]
                cv2.rectangle(frame, (0, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = y
            elif x-w < 0:
                zoomedframe = frame[y-h:y+h , 0:x+w] 
                cv2.rectangle(frame, (0, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = x
                warpy = h
            elif y-h < 0:
                zoomedframe = frame[0:y+h , x-w:x+w]
                cv2.rectangle(frame, (x-w, 0), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = y
            else:
                zoomedframe = frame[y-h:y+h , x-w:x+w] 
                cv2.rectangle(frame, (x-w, y-h), (x + w, y + h), (0, 255, 0), 2)
                warpx = w
                warpy = h   
            
            # step 2: zoom into the zoomed frame a certain zoom amount
            rot_mat = cv2.getRotationMatrix2D((warpx,warpy), angle, self.zoomscale)
            zoomedframe = cv2.warpAffine(zoomedframe, rot_mat, zoomedframe.shape[1::-1], flags=cv2.INTER_LINEAR)

            #step 3: replace the original cropped frame with the new zoomed in cropped frame
            if y-h < 0 and x-w < 0:
                frame[0:y+h , 0:x+w] =  zoomedframe
            elif x-w < 0:
                frame[y-h:y+h , 0:x+w] =  zoomedframe
            elif y-h < 0:
                frame[0:y+h , x-w:x+w] =  zoomedframe
            else:
                frame[y-h:y+h , x-w:x+w] =  zoomedframe


        
        return frame

    def closeEvent(self, event):
        """
        called when x button is pressed
        """
        
        if self.tracker is not None:
            self.tracker.stop()
        #self.recorder.stop()

        self.arduino.close()